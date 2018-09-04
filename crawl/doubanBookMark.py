#-*- coding: UTF-8 -*-

import time
import urllib
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook #Openpyxl进行excel2007的数据处理


#Some User Agents
hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
{'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


def book_spider(book_tag):
    '''
    :param book_tag: 豆瓣书签
    :return: 图书列表
    '''
    page_num=0;
    book_list=[]
    try_times=0
    
    while(1):
        #url='http://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/book?start=0' # For Test
        # https://book.douban.com
        # url=  'http://www.douban.com/tag/'   + urllib.request.quote(book_tag) +'/book?start='+str(page_num*15)
        url = 'https://book.douban.com/tag/' + urllib.request.quote(book_tag) # quote 将url中的特殊字符或汉字encode成指定编码
        time.sleep(np.random.rand()*5)
        
        try:
            # page_num%len(hds) 通过求余来依次循环
            req = urllib.request.Request(url, headers=hds[page_num%len(hds)])
            source_code = urllib.request.urlopen(req).read()
            plain_text=str(source_code)   
        except ( urllib.error.HTTPError, urllib.error.URLError) as e:
            print( e )
            continue

        # soup = BeautifulSoup(plain_text,from_encoding="utf-8")
        #soup = BeautifulSoup(plain_text,from_encoding="gb18030")
        soup = BeautifulSoup(plain_text)
        list_soup = soup.find('ul', {'class': 'subject-list'})
        
        try_times+=1;
        if list_soup==None and try_times<200:
            continue
        elif list_soup==None or len(list_soup)<=1:
            break # 200次请求之后，还无法获取数据，则退出
        
        for book_info in list_soup.findAll('li'):
            title = book_info.find('h2', {'class':''}).a['title']
            desc = book_info.find('div', {'class':'pub'}).string.strip()
            desc_list = desc.split('/')
            #book_url = book_info.find('p').string
            
            try:
                author_info = '作者/译者： ' + '/'.join(desc_list[0:1])
            except:
                author_info ='作者/译者： 暂无'
            try:
                pub_info = '出版信息： ' + '/'.join(desc_list[1])
            except:
                pub_info = '出版信息： 暂无'
            try:
                rating = book_info.find('span', {'class':'rating_nums'}).string.strip()
            except:
                rating='0.0'
            try:
                people_num = book_info.find('span', {'class':'pl'}).string.strip()
                # people_num = get_people_num(book_url)
                people_num = people_num[1:-1].strip('人评价)')
            except:
                people_num ='0'
            
            book_list.append([title,rating,people_num,author_info,pub_info])
            try_times=0 #set 0 when got valid information
        page_num+=1
        print( 'Downloading Information From Page %d' % page_num)
    return book_list


def get_people_num(url):
    #url='http://book.douban.com/subject/6082808/?from=tag_all' # For Test
    try:
        req = urllib.request.Request(url, headers=hds[np.random.randint(0,len(hds))])
        source_code = urllib.request.urlopen(req).read()
        plain_text=str(source_code)   
    except (urllib.error.HTTPError, urllib.error.URLError) as e:
        print("get_people_num" + e )
    soup = BeautifulSoup(plain_text)
    people_num=soup.find('div',{'class':'rating_sum'}).findAll('span')[1].string.strip()
    return people_num


def do_spider(book_tag_lists):
    book_lists=[]
    for book_tag in book_tag_lists:
        book_list=book_spider(book_tag)
        book_list=sorted(book_list,key=lambda x:x[1],reverse=True)
        book_lists.append(book_list)
    return book_lists


def print_book_lists_excel(book_lists,book_tag_lists):
    #wb=Workbook(optimized_write=True)
    wb = Workbook()
    ws=[]
    for i in range(len(book_tag_lists)):
        #ws.append(wb.create_sheet(title=book_tag_lists[i].decode())) #utf8->unicode
        ws.append( wb.create_sheet(title=book_tag_lists[i]) )
    for i in range(len(book_tag_lists)): 
        ws[i].append(['序号','书名','评分','评价人数','作者','出版社'])
        count=1
        for bl in book_lists[i]:
            ws[i].append([count,bl[0],float(bl[1]),int(bl[2]),bl[3],bl[4]])
            count+=1
    save_path='e://book_list'
    for i in range(len(book_tag_lists)):
        #save_path+=('-'+book_tag_lists[i].decode())
        save_path += ('-' + book_tag_lists[i])
    save_path+='.xlsx'
    wb.save(save_path)




if __name__=='__main__':
    
    book_tag_lists = ['模式识别']
    #book_tag_lists = ['python']
    book_lists = do_spider(book_tag_lists)
    print_book_lists_excel(book_lists,book_tag_lists)
    
    
